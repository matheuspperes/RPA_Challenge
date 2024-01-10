from openpyxl import load_workbook
from pathlib import Path


file_path = Path("Challenge 1", "challenge.xlsx")

def read_excel():
    wb = load_workbook(file_path)
    ws = wb.active
    rows = ws.max_row

    cases = []
    for row in range(2, rows -1):
        first_name = ws[f'A{row}'].value
        last_name = ws[f'B{row}'].value
        company = ws[f'C{row}'].value
        role = ws[f'D{row}'].value
        address = ws[f'E{row}'].value
        email = ws[f'F{row}'].value
        phone = ws[f'G{row}'].value

        current_case = {}
        current_case = {
            "first_name" : first_name,
            "last_name" : last_name,
            "company" : company,
            "role" : role,
            "address" : address,
            "email" : email,
            "phone" : phone
        }

        cases.append(current_case)